#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/Users/evhesap/Desktop/Kutuphane_calisiyor_AsilCalisma_AntiGravity_Org/kitap listesi.xlsx')
ws = wb.active

print(f"Excel dosyası: {ws.max_row} satır, {ws.max_column} sütun\n")
print("İLK 10 SATIR:\n")

# Header
header = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
print("HEADER:", header)
print("-" * 100)

# İlk 10 veri satırı
for row in range(2, min(12, ws.max_row + 1)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row, col).value
        row_data.append(str(cell_value) if cell_value is not None else "EMPTY")
    
    title = row_data[0] if len(row_data) > 0 else "?"
    author = row_data[1] if len(row_data) > 1 else "?"
    page = row_data[7] if len(row_data) > 7 else "?"
    
    print(f"Satır {row}: Başlık='{title}', Yazar='{author}', Sayfa={page}")
    if row <= 3:
        print(f"  Tüm sütunlar: {row_data[:10]}")

print("\n60 duplicate kontrolü için - Title+Author kombinasyonları:")
title_author_pairs = []
for row in range(2, ws.max_row + 1):
    title = ws.cell(row, 1).value
    author = ws.cell(row, 2).value
    if title and author:
        pair = f"{str(title).strip()} | {str(author).strip()}"
        title_author_pairs.append(pair)

# Duplicate sayısı
from collections import Counter
counts = Counter(title_author_pairs)
duplicates = {k: v for k, v in counts.items() if v > 1}

print(f"\nToplam kitap: {len(title_author_pairs)}")
print(f"Duplicate çiftler: {len(duplicates)}")
print(f"Toplam duplicate satır: {sum(v - 1 for v in duplicates.values())}")

if duplicates:
    print("\nİlk 5 duplicate:")
    for i, (pair, count) in enumerate(list(duplicates.items())[:5]):
        print(f"  {pair} -> {count} kez")
