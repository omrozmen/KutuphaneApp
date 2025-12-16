#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/Users/evhesap/Desktop/Kutuphane_calisiyor_AsilCalisma_AntiGravity_Org/kitap listesi.xlsx')
ws = wb.active

print(f"Excel dosyası analizi:")
print(f"Toplam satır: {ws.max_row} (header dahil)")
print(f"Sütun sayısı: {ws.max_column}\n")

# Header göster
header = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
print(f"Sütunlar: {header}\n")

print("=" * 100)
print("BOŞ BAŞLIK veya YAZAR İÇEREN SATIRLAR:")
print("=" * 100)

empty_rows = []
valid_rows = 0

for row in range(2, ws.max_row + 1):
    title_cell = ws.cell(row, 1)  # Başlık (A sütunu)
    author_cell = ws.cell(row, 2)  # Yazar (B sütunu)
    
    title_value = title_cell.value
    author_value = author_cell.value
    
    # Değerleri kontrol et
    title_str = str(title_value).strip() if title_value is not None else ""
    author_str = str(author_value).strip() if author_value is not None else ""
    
    # Boş satırları tespit et
    if not title_str or not author_str:
        empty_rows.append({
            'row': row,
            'title': repr(title_value),  # repr() ile tam değeri göster
            'author': repr(author_value),
            'title_type': type(title_value).__name__,
            'author_type': type(author_value).__name__
        })
        
        # İlk 20 boş satırı detaylı göster
        if len(empty_rows) <= 20:
            print(f"\nSatır {row}:")
            print(f"  Başlık: {repr(title_value)} (Tip: {type(title_value).__name__})")
            print(f"  Yazar:  {repr(author_value)} (Tip: {type(author_value).__name__})")
            
            # Tüm sütunları göster
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):
                val = ws.cell(row, col).value
                row_data.append(repr(val)[:30])
            print(f"  İlk 5 sütun: {row_data}")
    else:
        valid_rows += 1

print("\n" + "=" * 100)
print(f"ÖZET:")
print(f"  Geçerli satırlar (Başlık VE Yazar dolu): {valid_rows}")
print(f"  Boş satırlar (Başlık VEYA Yazar boş): {len(empty_rows)}")
print(f"  Toplam veri satırı: {ws.max_row - 1}")
print("=" * 100)

if len(empty_rows) > 20:
    print(f"\n⚠️  Toplam {len(empty_rows)} boş satır var ama sadece ilk 20'sini gösterdim.")
    print(f"Boş satır numaraları: {[r['row'] for r in empty_rows]}")
